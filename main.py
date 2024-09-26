"""
  The following program provides the user with a search interface for certificates pertaining to the 
  computer science path. It provides four different search parameters in order to match the user with 
  a suitable certificate from the Certification_Data.xlsx database. The first search parameter is a 
  dropdown selecting the specific industry within Computer Science that the user is looking to delve 
  into (e.g. Cybersecurity). The second search parameter is a dropdown selecting which organization 
  the user would like to earn a certificate from; the user can select a specific option such as 
  Microsoft, or they can select the option 'Any.' The third search parameter is a combobox dropdown 
  where the user can select the maximum cost of the certificate, or they can type the maximum cost in 
  themselves. The final search parameter is the selection of the location (online/in-person/either). 
  After the user has entered their search parameters, they click the search button, and the outputted 
  certificates that match are displayed. Along with the outputted names of each certificate, the 
  cost, location, and the link leading to the registration page are provided. If no matches are given 
  or an error is shown, the user can reset and alter their search parameters.
"""

import customtkinter  # This library provides custom widgets and functionality for Tkinter GUI toolkit.
# Citation for customtkinter library:
# Author: Tom Schimansky
# Title: CustomTkinter
# Year: 2024
# Version: 5.2.2
# URL or Source: [GitHub Repository](https://github.com/TomSchimansky/CustomTkinter)

from openpyxl.workbook import Workbook  # This library is used to create a new Excel workbook.
# Citation for openpyxl library:
# Authors: Eric Gazoni, Charlie Clark, and contributors
# Title: openpyxl
# Year: 2021
# Version: 3.0.9
# URL or Source: https://openpyxl.readthedocs.io/en/stable/

from openpyxl import load_workbook  # This library is used to load an existing Excel workbook.
# Citation for openpyxl library (same as above):
# Authors: Eric Gazoni, Charlie Clark, and contributors
# Title: openpyxl
# Year: 2021
# Version: 3.0.9
# URL or Source: https://openpyxl.readthedocs.io/en/stable/

import webbrowser
# Citation for webbrowser module:
# Title: webbrowser
# Year: Not applicable (Included in Python standard library)
# URL or Source: https://docs.python.org/3/library/webbrowser.html

# Create workbook instance
wb = Workbook()

# Load Certification Data
wb = load_workbook('Certification_Data.xlsx')

# Create active worksheet
ws = wb.active

# Get column data
# Assigns each column to respective variable name
names_col = ws['A']
industries_col = ws['B']
organizations_col = ws['C']
costs_col = ws['D']
locations_col = ws['E']
links_col = ws['F']

# Establishes industries list
industries = []

# Establishes organizations list
organizations = []

# Adds each industry from excel database into the industry list
for industry in industries_col[1:]:
    if industry.value != None:
        if industry.value not in industries:
            industries.append(industry.value)

# Adds each organization from the excel database into the organization list
for organization in organizations_col[1:]:
    if organization.value != None:
        if organization.value not in organizations:
            organizations.append(organization.value)    
# Adds the 'All' option for the user
organizations[0:0] = ["All"]

# Establishes the selection list which will hold the inputted user values
selection = ["Industry", "Organization", "Cost", "Location"]

# Used customtkinter documentation to create searchMenuFrame Class
class searchMenuFrame(customtkinter.CTkFrame):
    def __init__(self, master):
        super().__init__(master)

        # Establishes number of rows and columns in the frame
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(8, weight = 1)

        """
        This function resets the to the search menu screen.

        Precondition: The output screen, whether this is the outputted matched certificate, a no match certificate message, or an error message.
        Postcondition: Search screen appears and user can input new search parameters.
        """
        def reset():

            # Deletes everything on the screen
            for widget in self.winfo_children():
                widget.destroy()

            # Calls for the search screen to appear
            searchScreen()

        """
        This function opens a website url link using the webbrowser python library

        Precondition: The output screen, whether this is the outputted matched certificate, a no match certificate message, or an error message.
        Postcondition: A new tab in the users default search engine showing the website where the certificate can be found.
        """
        # Opens website based on url given
        def open_website(url):
            # Uses pythons webbrowser library to open the url in a new tab
            webbrowser.open_new(url)

        """
        This function is responsible for displaying the respective output screen based on the user's search inquiry

        Precondition: The search screen where the user presses the search button.
        Postcondition: The output screen. This could be one of three things: the name of the certificate that matches the user's search along with it's cost, location and link, a no match certificate message, or an error message.
        """
        # Outputs the certificates that match the user's needs
        def searchOutput(matched):

            # Deletes current screen
            for widget in self.winfo_children():
                widget.destroy()

            # Creates Reset Button
            self.button_reset = customtkinter.CTkButton(self, corner_radius=10, text="Reset", command=reset)
            self.button_reset.grid(row=8, column=0, pady=(20,10), sticky="s")

            # If error in searching process then output the error message
            if matched == "Error":
                self.error_text = customtkinter.CTkLabel(self, text="Error: Cost must be\nnumerical value\n\nPlease Reset", font=("Cascadia Code", 15))    
                self.error_text.grid(row = 0, column = 0, pady = (5,0))
                return False

            # If no certificates matched in search process then output message
            if matched == []:
                self.nomatches_text = customtkinter.CTkLabel(self, text="Sorry, no certifcation\nmatches your request.\n\nPlease Reset to\nSearch Again", font=("Cascadia Code", 15))    
                self.nomatches_text.grid(row = 0, column = 0, pady = (5,0))
                return False

            # Iterates through the matched certificates
            # matched is a list containing each row number of the matched certificate from the excel database
            for i, row in enumerate(matched):

                # Collects the certificates name and url for respective certificate
                cert_name = names_col[row-1].value
                url = links_col[row - 1].value

                # First Outputs Name of Certificate in Bold
                self.certificate_title = customtkinter.CTkLabel(self, text=cert_name, font=("Cascadia Code", 15, "bold"))    
                self.certificate_title.grid(row = i, column = 0, padx = (15,0), pady = (5,0), sticky='nw')

                # Creates Bulleted List of the cost, location, and link to certificate
                # Cost value outputted
                self.certificate_cost = customtkinter.CTkLabel(self, text= ("- Cost: " + str(costs_col[row-1].value)), font=("Cascadia Code", 15))    
                self.certificate_cost.grid(row = i, column = 0, padx = (30,0), pady = (27,0), sticky='nw')

                # Location (In-person / Online / Both) outputted
                self.certificate_location = customtkinter.CTkLabel(self, text= ("- Location: " + locations_col[row-1].value), font=("Cascadia Code", 15))    
                self.certificate_location.grid(row = i, column = 0, padx = (30,0), pady = (49,0), sticky='nw')

                # Link outputted as a button
                # Button calls back to openwebsite function
                self.certificate_link = customtkinter.CTkButton(self, text="Link", font=("Cascadia Code", 15), command=lambda u=url: open_website(u))    
                self.certificate_link.grid(row = i, column = 0, padx = (15,0), pady = (80,0), sticky='sw')  

        """
        This function takes the user's inputs and finds any certificates from the Certification_Data.xlsx database that match the user's request.

        Precondition: The user selects the industry, organization, location, and maximum cost of their desired certificate and then presses search. The user's request is compiled and sent to this function as a paramter.
        Postcondition: searchOutput function is called passing each row of each certificate from the database that matches the user's search request in a list known as 'matched'.
        """
        # Searches through data base and finds each certificate that matches the search parameters
        def searchCertificate(selection):

            # Converts cost keywords (from combobox dropdown) to numerical values
            if selection[2] == "Free":
                selection[2] = "0"
            if selection[2] == "Any":
                # Converts 'Any' to the max value of all certificates
                selection[2] = "2000"

            # Try block ensures user did not enter a non-numerical value
            try:
                # Converts cost selection to numerical value
                selection[2] = float(selection[2])
            except:
                # Calls search output with 'Error' as arg so user will be notified
                searchOutput("Error")
                return

            # Establishes list that will house the row numbers for each matched certificate from the excel database
            matched = []

            # Iterates through each industry that matches
            for industry in industries_col:
                if selection[0] == industry.value:

                    # For each matched industry, find the matched organization
                    for organization in organizations_col:
                        if selection[1] == "All" or selection[1] == organization.value:

                            # For each matched organization and industry find the matched cost
                            for cost in costs_col[1:]:
                                    # Ensures the cell has a value
                                    if cost.value is not None:
                                        # Checks to see if the user selection is greater the actual cost
                                        if selection[2] >= float(cost.value):

                                            # Finds the matched location
                                            for location in locations_col:
                                                if selection[3] == location.value or selection[3] == "Either" or location.value == "Both":

                                                    # Compares all the row values to see if they match
                                                    if selection[1] == "All" and (industry.row == cost.row == location.row) or (industry.row == cost.row == location.row == organization.row):

                                                        # Makes sure the matched row was not already found
                                                        if industry.row not in matched:

                                                            # Adds the database row that effectively suits the users search parameters
                                                            matched.append(industry.row)      

            # Outputs the matched data                                  
            searchOutput(matched)

        """
        This function is responsible for gathering the user's inputted values from each dropdown/combobox and storing it in a parsable datastructure ('selection'). The function then calls the function searchCertificate passing this datastructure as the parameter.

        Precondition: The user selects the industry, organization, location, and maximum cost of their desired certificate and then presses search.
        Postcondition: The backend calls the function searchCertificate using this compiled data to effectively find a suitable certificate.
        """
        # Gathers and compiles users request into selection list
        def searchDataCompile():

            # Uses customtkinter get() function to get each dropdown value
            # Assigns each value to specified selection list indices
            selection[0] = self.dropdown_industry.get()
            selection[1] = self.dropdown_organization.get()
            selection[2] = self.combobox_cost.get()
            selection[3] = self.dropdown_location.get()

            # Searchs the excel database to find a certificate the matches the user search parameters
            searchCertificate(selection)

        """
        This function is responsible for creating the search screen.

        Precondition: The screen is empty or the reset button has been clicked from the output screen.
        Postcondition: The search screen appears. This includes a dropdown for industry, organization, and location selection; a combobox dropdown for the maximum cost selection; a search button.
        """
        # Creates dropdowns for each input
        def searchScreen():

            # Industry Dropdown Title
            self.dropdown_industry_title = customtkinter.CTkLabel(self, text="Industry", font=("Cascadia Code", 15))    
            self.dropdown_industry_title.grid(row = 0, column = 0, pady = (5,0))
            # Industry Dropdown Selection Created
            self.dropdown_industry = customtkinter.CTkOptionMenu(self, values=industries)
            self.dropdown_industry.grid(row=1, column=0)

            # Organization Dropdown Title
            self.dropdown_organization_title = customtkinter.CTkLabel(self, text="Organization", font=("Cascadia Code", 15))    
            self.dropdown_organization_title.grid(row = 2, column = 0, pady = (5,0))
            # Organization Dropdown Selection Created
            self.dropdown_organization = customtkinter.CTkOptionMenu(self, values=organizations)
            self.dropdown_organization.grid(row=3, column=0)

            # Cost Input Title
            self.combobox_cost_title = customtkinter.CTkLabel(self, text="Maximum Cost (USD)", font=("Cascadia Code", 15))    
            self.combobox_cost_title.grid(row = 4, column = 0, pady = (5,0))
            # Cost Combobox Created
            combobox_cost_var = customtkinter.StringVar(value="Any")
            self.combobox_cost = customtkinter.CTkComboBox(self, values=["Free", "100", "250", "500", "750", "1000", "2000"], variable=combobox_cost_var)
            self.combobox_cost.grid(row=5, column=0)

            # Location Dropdown Title
            self.dropdown_location_title = customtkinter.CTkLabel(self, text="Location", font=("Cascadia Code", 15))    
            self.dropdown_location_title.grid(row = 6, column = 0, pady = (5,0))
            # Location Dropdown Selection Created
            self.dropdown_location = customtkinter.CTkOptionMenu(self, values=["Online", "In-Person", "Either"])
            self.dropdown_location.grid(row=7, column=0)

            # Search Button
            # Initiates searchDataCompile Function
            self.button_search = customtkinter.CTkButton(self, corner_radius=10, text="Search", command=searchDataCompile)
            self.button_search.grid(row=8, column=0, pady=(20,10), sticky="s")

        # Creates the search screen
        searchScreen()

# Used customtkinter documentation to create App Class
class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # Creates App Window
        self.title("CS Certificate Finder")
        self.geometry("400x450")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        # Title of the app established
        self.heading = customtkinter.CTkLabel(self, text="Computer Science\nCertificate Finder", font=("Cascadia Code", 25))
        self.heading.grid(row=0, column=0, pady=(10, 25))

        # Creates the search menu frame by referencing the searchMenuFrame Class
        self.search_menu_frame = searchMenuFrame(self)
        self.search_menu_frame.grid(row=1, column=0, padx=25, pady=(0, 25), sticky="nsew")


# Runs App
app = App()
app.mainloop()